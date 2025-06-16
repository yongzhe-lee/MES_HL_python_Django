from domain.services.logging import LogWriter
from domain.services.sql import DbUtil
from domain.services.common import CommonUtil
#from django.db import transaction
from domain.models.definition import Material
from domain.services.definition.material import MaterialService

def material(context):
    '''
    /api/definition/material
    
    작성명 : 품목정보
    작성자 : 
    작성일 : 
    비고 :

    -수정사항-
    수정일             작업자     수정내용

    '''
    gparam = context.gparam
    posparam = context.posparam
    request = context.request

    material_service = MaterialService()
    
    action = gparam.get('action', 'read')
    source = f'/api/definition/material?action={action}'
    result = {}
    try:
        if action =='read':
            factory = gparam.get('srch_factory')
            mat_type= gparam.get('mat_type')
            keyword = gparam.get('keyword')

            dc = {}
            dc['factory'] = factory
            dc['mat_type'] = mat_type
            dc['keyword'] = keyword

            sql = '''
                SELECT 
                m.id AS material_id
                , m."Factory_id" AS factory_id
                , f."Code" AS factory_code
                , f."Name" AS factory_name
                , m."Code" AS material_code
                , m."Name" AS material_name
                , m.bare_pcb_cd
                , m.panel_mag
                , m.pcb_array
                , m.sap_top_bottom
                , m.solder_paste
                , m.sub_pcb_cd
                , m.top_bottom
                , m.use_yn
                , m."mat_type"
                , m."Standard" AS standard
                , mg."Name" AS mat_grp_nm
                , m.mat_type
                , m."BasicUnit" AS basic_unit
                , m."CycleTime" AS cycle_time
                , m."in_price" AS in_price
                , m."out_price" AS out_price
                , m.supplier_pk as supplier
            FROM  material m
                left join mat_grp mg on mg.id=m.mat_grp_id 
                INNER JOIN factory f on m."Factory_id" = f.id
            WHERE 1=1
            '''
            if factory:
                sql += ''' 
                AND f.id = %(factory)s
                '''
            if mat_type:
                sql += ''' 
                AND f.id = %(mat_type)s
                '''
            if keyword:
                sql += ''' 
                AND (
                    UPPER(m."Name") LIKE CONCAT('%%', UPPER(%(keyword)s), '%%')
                    OR UPPER(m."Code") LIKE CONCAT('%%', UPPER(%(keyword)s), '%%')
                    OR UPPER(m."ItemGroup") LIKE CONCAT('%%', UPPER(%(keyword)s), '%%')
                    OR UPPER(m."ItemType") LIKE CONCAT('%%', UPPER(%(keyword)s), '%%')
                ) 
                '''

            sql += '''
            ORDER BY m."Name"
            '''
            items = DbUtil.get_rows(sql, dc)
            result['success'] = True
            result['items'] = items

        if action =='read_modal':

            keyword = gparam.get('keyword')
            ItemType = gparam.get('ItemType')   
            Supplier = gparam.get('Supplier')

            result = material_service.get_material_modal(keyword, ItemType, Supplier)        

        elif action == 'detail':
            id = gparam.get('id')
            dc = {"id" : id}

            sql = '''
            select
            m.id AS material_id
            , m."Factory_id" AS factory_id
            , f."Code" AS factory_code
            , f."Name" AS factory_name
            , m."Code" AS material_code
            , m."Name" AS material_name
            , m."mat_type"
            , m."Standard" AS standard
            , m."BasicUnit" AS basic_unit
            , m."CycleTime" AS cycle_time
            , m."in_price" AS in_price
            , m."out_price" AS out_price
            FROM  material m
            inner join factory f on m."Factory_id" = f.id
            where 
            m.id = %(id)s
            '''
            
            data = DbUtil.get_row(sql, dc)
            result['success'] = True
            result['data'] = data

        elif action == 'save':
            material_id = CommonUtil.try_int(posparam.get('material_id'))
            material = Material.objects.get(id = material_id)

            bar_pcb_cd = posparam.get('bar_pcb_cd')
            sub_pcb_cd = posparam.get('sub_pcb_cd')
            pcb_array = posparam.get('pcb_array')
            panel_mag = posparam.get('panel_mag')
            top_bottom = posparam.get('top_bottom')
            solder_paste = posparam.get('solder_paste')
            sap_top_bottom = posparam.get('sap_top_bottom')

            material.bare_pcb_cd = bar_pcb_cd
            material.sub_pcb_cd = sub_pcb_cd
            material.pcb_array = pcb_array
            material.panel_mag = panel_mag
            material.top_bottom = top_bottom
            material.solder_paste = solder_paste
            material.sap_top_bottom = sap_top_bottom

            material.set_audit(request.user)
            material.save()

            result = {'success' : True, 'material_id' : material_id}

        elif action == 'delete':
            id = posparam.get('material_id')

            if id:
                material = Material.objects.filter(id = id).first()
                material.delete()

            result = {'success' : True}

        elif action=="smt_mat_info_excel":
            import openpyxl

            mat_file = "c:\\temp\\SMT 제품 정보.xlsx"
            workbook = openpyxl.load_workbook(mat_file)
            sheet = workbook['제품']

            dic_mat_grp = {}

            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
                mat_cd = row[0].value
                mat_nm = row[1].value
                mat_grp = row[2].value

                print(mat_cd)

                if mat_grp not in dic_mat_grp:
                    dic_mat_grp[mat_grp] = mat_grp



            for mg in dic_mat_grp:
                print(mg)

            workbook.close()
            

    except Exception as ex:
        LogWriter.add_dblog('error', source, ex)
        result = {'success':False, 'message':str(ex) }

    return result